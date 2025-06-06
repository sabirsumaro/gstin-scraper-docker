import streamlit as st
import openpyxl
import tempfile
import time
import pandas as pd
from io import BytesIO
from selenium import webdriver
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.common.by import By
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from webdriver_manager.chrome import ChromeDriverManager

# Pre-download ChromeDriver once
CHROMEDRIVER_EXECUTABLE_PATH = ChromeDriverManager().install()

def to_excel_bytes(df):
    output = BytesIO()
    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        df.to_excel(writer, index=False)
    return output.getvalue()

def setup_driver():
    chrome_options = Options()
    chrome_options.add_argument("--headless")
    chrome_options.add_argument("--disable-gpu")
    chrome_options.add_argument("--no-sandbox")
    chrome_options.add_argument("--disable-dev-shm-usage")
    service = Service(CHROMEDRIVER_EXECUTABLE_PATH)
    return webdriver.Chrome(service=service, options=chrome_options)

def process_gstin_list(gstin_list):
    driver = setup_driver()
    results = []

    def get_data(label):
        try:
            element = driver.find_element(By.XPATH, f"//strong[contains(text(), '{label}')]")
            return element.find_element(By.XPATH, "..").text.replace(label, "").strip()
        except:
            return ""

    for i, gstin in enumerate(gstin_list):
        result = [gstin, "", "", "", "", "", ""]

        try:
            driver.get("https://irisgst.com/irisperidot/")
            WebDriverWait(driver, 15).until(EC.presence_of_element_located((By.ID, "gstinno")))
            driver.execute_script("document.querySelectorAll('.popup, iframe').forEach(p => p.remove())")
            input_box = driver.find_element(By.ID, "gstinno")
            input_box.clear()
            input_box.send_keys(gstin)
            driver.find_element(By.XPATH, "//button[contains(text(), 'SEARCH')]").click()

            start_time = time.time()
            while time.time() - start_time < 20:
                if "Trade Name -" in driver.page_source:
                    break
                time.sleep(1)

            driver.execute_script("document.querySelectorAll('.popup, iframe').forEach(p => p.remove())")

            result = [
                gstin,
                get_data("Trade Name -"),
                get_data("Legal Name of Business -"),
                get_data("Principal Place of Business -"),
                get_data("Additional Place of Business -"),
                get_data("State Jurisdiction -"),
                "Success"
            ]
        except Exception as e:
            result[-1] = f"Error: {str(e)}"

        results.append(result)
        st.progress(int((i + 1) / len(gstin_list) * 100))
        st.text(f"Processed {i + 1} of {len(gstin_list)}")

    driver.quit()
    return results

st.set_page_config(page_title="GSTIN Scraper", layout="centered")
st.title("GSTIN Bulk Scraper 🔍 (Single Chrome - Stable)")
st.markdown("✅ Safe for Render — processes GSTINs one by one in a single Chrome instance.")

sample_data = pd.DataFrame({'GSTIN': ['06ABCDE1234F1Z5', '07XYZAB1234L1Z2']})
with st.expander("📥 Download Sample Template"):
    excel_bytes = to_excel_bytes(sample_data)
    st.download_button("📥 Download Excel Template", data=excel_bytes, file_name="Bulk_GSTIN_Input_Template.xlsx")

uploaded_file = st.file_uploader("📤 Upload your GSTIN Excel file", type=["xlsx"])
if uploaded_file is not None:
    with tempfile.NamedTemporaryFile(delete=False, suffix=".xlsx") as tmp:
        tmp.write(uploaded_file.read())
        tmp_path = tmp.name

    wb = openpyxl.load_workbook(tmp_path)
    ws = wb.active
    gstin_list = [row[0].value for row in ws.iter_rows(min_row=3, min_col=1, max_col=1) if row[0].value]

    st.info(f"Total GSTINs to process: {len(gstin_list)}")
    results = process_gstin_list(gstin_list)

    # Write to Excel
    output_wb = openpyxl.Workbook()
    output_ws = output_wb.active
    output_ws.title = "Results"
    output_ws.append(["GSTIN", "Trade Name", "Legal Name", "Principal Place", "Additional Place", "Jurisdiction", "Status"])
    for row in results:
        output_ws.append(row)
    output_final = tmp_path.replace(".xlsx", "_Result.xlsx")
    output_wb.save(output_final)

    with open(output_final, "rb") as f:
        st.success("✅ Done! Click below to download your result:")
        st.download_button("📄 Download Result Excel", f, file_name="Bulk_GSTIN_Result.xlsx")
